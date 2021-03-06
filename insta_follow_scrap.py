from InstagramAPI import InstagramAPI
import tkinter as tk
from tkinter import filedialog
import openpyxl
import xlrd
from tkinter import messagebox
import threading
import os
import time

api = None


def login(username, password):
    global api
    api = InstagramAPI(username, password)
    return api.login()


def logout():
    global api
    if api.isLoggedIn:
        api.logout()


class StoppableThread(threading.Thread):
    def __init__(self, target):
        super(StoppableThread, self).__init__(target=target)
        self._stop_event = threading.Event()
        # self.is_stopped = False

    def stop(self):
        self._stop_event.set()
        # self.is_stopped = True

    def stopped(self):
        return self._stop_event.is_set()


class MainView:
    def __init__(self, master):
        self.master = master
        master.title("Scrap instagram")

        master.grid_columnconfigure(0, weight=1)
        master.grid_rowconfigure(3, weight=1)

        frame1 = tk.Frame(master, pady=3)
        frame2 = tk.Frame(master, pady=3)
        frame3 = tk.Frame(master, pady=3)
        frame4 = tk.Frame(master, pady=3)
        frame5 = tk.Frame(master, pady=3)

        frame1.grid(row=0, sticky='ew')
        frame2.grid(row=1, sticky='ew')
        frame3.grid(row=2, sticky='ew')
        frame4.grid(row=3, sticky='nsew')
        frame5.grid(row=4, sticky='ew')

        tk.Label(frame1, text='Select excel file : ', width=15, anchor='e').pack(side='left')
        self.excel_file = tk.StringVar()
        tk.Entry(frame1, width=50, textvariable=self.excel_file).pack(side='left', expand=True)

        tk.Label(frame2, text='Username : ', width=15, anchor='e').pack(side='left')
        self.username = tk.StringVar()
        tk.Entry(frame2, width=50, textvariable=self.username).pack(side='left', expand=True)

        tk.Label(frame3, text='Password : ', width=15, anchor='e').pack(side='left')
        self.password = tk.StringVar()
        tk.Entry(frame3, width=50, textvariable=self.password).pack(side='left', expand=True)

        browse1_button = tk.Button(frame1, text="Browse", width=8, command=self.browse_excel_file)
        browse1_button.pack(side='right')

        self.progress = tk.StringVar()
        label5 = tk.Label(frame4, textvariable=self.progress, borderwidth=2, relief='groove')
        label5.pack(side='left', expand=True, fill='both')

        self.button_text = tk.StringVar()
        self.button_text.set("Start")
        toggle_button = tk.Button(frame5, textvariable=self.button_text, width=8, command=self.toggle)
        toggle_button.pack()

        self.thread = None
        self.users = []
        self.target_user = {}
        self.follower_limit = 0

    def browse_excel_file(self):
        text = filedialog.askopenfilename(title='Select input excel file', filetypes=[("Excel files", "*.xlsx")])
        if text != '':
            self.excel_file.set(text)

    def toggle(self):
        if self.thread is None:
            if self.excel_file.get() == '':
                messagebox.showerror('Error', 'Please input proper paths')
                return
            self.users = []
            self.target_user = {
                'name': '',
                'status': 'pending'
            }
            self.follower_limit = 0

            if not os.path.exists(self.excel_file.get()):
                return
            xls = xlrd.open_workbook(self.excel_file.get())

            source_sheet = xls.sheet_by_index(0)

            for row in range(1, source_sheet.nrows):
                self.users.append({'name': str(source_sheet.row(row)[0].value),
                                   'status': 'pending'})
            self.update_progress()

            if len(str(source_sheet.row(1)[1].value)) != 0:
                self.target_user['name'] = str(source_sheet.row(1)[1].value)
            if len(str(source_sheet.row(1)[2].value)) != 0:
                self.follower_limit = int(source_sheet.row(1)[2].value)

            if not login(self.username.get(), self.password.get()):
                messagebox.showerror('Error', 'Login failed')
                return
            self.thread = StoppableThread(self.scrap_data)
            self.thread.setDaemon(True)
            self.thread.start()
            self.button_text.set("Stop")
        else:
            self.thread.stop()
            while self.thread.isAlive():
                self.thread.join(1)
            self.thread = None
            self.button_text.set("Start")
            logout()
        pass

    def update_progress(self):
        message = ''
        for user in [self.target_user] + self.users:
            if user['status'] == 'pending':
                message += '%s is pending\n' % user['name']
            elif user['status'] == 'working follower':
                message += '-->working on %s followers(%d/%d)\n' % (user['name'], user['worked_count'], user['total_count'])
            elif user['status'] == 'working following':
                message += '-->working on %s followings(%d/%d)\n' % (user['name'], user['worked_count'], user['total_count'])
            elif user['status'] == 'error':
                message += 'error occurred on %s\n' % user['name']
            elif user['status'] == 'not found':
                message += '%s is not found\n' % user['name']
            elif user['status'] == 'completed':
                message += 'completed scraping on %s\n' % user['name']
        self.progress.set(message)

    def scrap_data(self):
        output_file = self.excel_file.get().replace('.xls', '_result.xls')
        book = openpyxl.Workbook()
        source_sheet = book.active
        source_sheet.title = 'source'
        follower_sheet = book.create_sheet('results')
        summary_sheet = book.create_sheet('summary')
        home_follower_sheet = book.create_sheet('home followers')
        home_following_sheet = book.create_sheet('home followings')
        home_follow_delta_sheet = book.create_sheet('home follow delta')
        all_followers = {}
        home_followers = []

        source_sheet.cell(row=1, column=1).value = 'target account'
        source_sheet.cell(row=1, column=2).value = 'exclude followers from'
        source_sheet.cell(row=1, column=3).value = 'exclude users with follower count less than'

        ind = 2
        for user in self.users:
            source_sheet.cell(row=ind, column=1).value = user['name']
            ind += 1
        follower_sheet.cell(row=1, column=1).value = 'target account'
        follower_sheet.cell(row=1, column=2).value = 'username'
        follower_sheet.cell(row=1, column=3).value = 'followers'
        follower_sheet.cell(row=1, column=4).value = 'followings'
        summary_sheet.cell(row=1, column=1).value = 'unique users from results'
        summary_sheet.cell(row=1, column=2).value = 'target following'
        summary_sheet.cell(row=1, column=3).value = 'follower count'
        summary_sheet.cell(row=1, column=4).value = 'following count'
        home_follower_sheet.cell(row=1, column=1).value = 'username'
        home_follower_sheet.cell(row=1, column=2).value = 'follower count'
        home_follower_sheet.cell(row=1, column=3).value = 'following count'
        home_following_sheet.cell(row=1, column=1).value = 'username'
        home_following_sheet.cell(row=1, column=2).value = 'follower count'
        home_following_sheet.cell(row=1, column=3).value = 'following count'
        home_follow_delta_sheet.cell(row=1, column=1).value = 'username'

        if self.follower_limit != 0:
            source_sheet.cell(row=2, column=3).value = self.follower_limit

        if len(self.target_user['name']) > 0:
            source_sheet.cell(row=2, column=2).value = self.target_user['name']
            while not api.searchUsername(self.target_user['name']) or 'user' not in api.LastJson:
                if api.LastJson['message'] == 'Please wait a few minutes before you try again.':
                    print('sleeping for a minute')
                    time.sleep(60)
                else:
                    self.target_user['status'] = 'not found'
                    self.update_progress()
                    break

            ind = 2
            if api.LastJson['status'] == 'ok':
                user_id = api.LastJson['user']['pk']
                follower_count = api.LastJson['user']['follower_count']
                following_count = api.LastJson['user']['following_count']
                max_id = ''

                self.target_user['status'] = 'working follower'
                self.target_user['total_count'] = follower_count
                self.target_user['worked_count'] = 0
                self.update_progress()

                while True:
                    if self.thread.stopped():
                        return
                    try:
                        api.getUserFollowers(str(user_id), max_id)
                        has_max_id = False
                        if api.LastJson['status'] != 'ok':
                            print('sleeping for a minute')
                            time.sleep(60)
                            continue

                        if 'next_max_id' in api.LastJson:
                            max_id = api.LastJson['next_max_id']
                            has_max_id = True

                        if len(api.LastJson['users']) == 0:
                            break
                        else:
                            for follower in api.LastJson['users']:
                                follower_name = follower['username']
                                home_followers.append(follower_name)

                                while not api.searchUsername(follower_name) or 'user' not in api.LastJson:
                                    if api.LastJson['message'] == 'Please wait a few minutes before you try again.':
                                        print('sleeping for a minute')
                                        time.sleep(60)
                                    else:
                                        break

                                if 'user' in api.LastJson:
                                    home_follower_sheet.cell(row=ind, column=1).value = follower_name
                                    home_follower_sheet.cell(row=ind, column=2).value = api.LastJson['user']['follower_count']
                                    home_follower_sheet.cell(row=ind, column=3).value = api.LastJson['user']['following_count']
                                    ind += 1

                                self.target_user['worked_count'] += 1
                                self.update_progress()

                        if not has_max_id:
                            break
                    except Exception as ex:
                        print(ex)
                        print('sleeping for a minute')
                        time.sleep(60)

                self.target_user['status'] = 'working following'
                self.target_user['total_count'] = following_count
                self.target_user['worked_count'] = 0
                self.update_progress()

                ind = 2
                delta_ind = 2
                max_id = ''

                while True:
                    if self.thread.stopped():
                        return
                    try:
                        api.getUserFollowings(str(user_id), max_id)
                        has_max_id = False
                        if api.LastJson['status'] != 'ok':
                            print('sleeping for a minute')
                            time.sleep(60)
                            continue

                        if 'next_max_id' in api.LastJson:
                            max_id = api.LastJson['next_max_id']
                            has_max_id = True

                        if len(api.LastJson['users']) == 0:
                            break
                        else:
                            for following in api.LastJson['users']:
                                following_name = following['username']

                                while not api.searchUsername(following_name) or 'user' not in api.LastJson:
                                    if api.LastJson['message'] == 'Please wait a few minutes before you try again.':
                                        print('sleeping for a minute')
                                        time.sleep(60)
                                    else:
                                        break

                                if 'user' in api.LastJson:
                                    home_following_sheet.cell(row=ind, column=1).value = following_name
                                    home_following_sheet.cell(row=ind, column=2).value = api.LastJson['user']['follower_count']
                                    home_following_sheet.cell(row=ind, column=3).value = api.LastJson['user']['following_count']
                                    ind += 1

                                if following_name not in home_followers:
                                    home_follow_delta_sheet.cell(row=delta_ind, column=1).value = following_name
                                    delta_ind += 1

                                self.target_user['worked_count'] += 1
                                self.update_progress()

                        if not has_max_id:
                            break
                    except Exception as ex:
                        print(ex)
                        print('sleeping for a minute')
                        time.sleep(60)

            if 'working' in self.target_user['status']:
                self.target_user['status'] = 'completed'
                self.update_progress()

        ind = 2
        for user in self.users:
            if self.thread.stopped():
                return

            while not api.searchUsername(user['name']) or 'user' not in api.LastJson:
                if api.LastJson['message'] == 'Please wait a few minutes before you try again.':
                    print('sleeping for a minute')
                    time.sleep(60)
                else:
                    user['status'] = 'not found'
                    self.update_progress()
                    break

            if api.LastJson['status'] == 'ok':
                user_id = api.LastJson['user']['pk']
                follower_count = api.LastJson['user']['follower_count']
                max_id = ''

                user['status'] = 'working follower'
                user['total_count'] = follower_count
                user['worked_count'] = 0
                self.update_progress()
                while True:
                    if self.thread.stopped():
                        return
                    try:
                        api.getUserFollowers(str(user_id), max_id)
                        has_max_id = False
                        if api.LastJson['status'] != 'ok':
                            print('sleeping for a minute')
                            time.sleep(60)
                            continue

                        if 'next_max_id' in api.LastJson:
                            max_id = api.LastJson['next_max_id']
                            has_max_id = True

                        if len(api.LastJson['users']) == 0:
                            break
                        else:
                            for follower in api.LastJson['users']:
                                follower_name = follower['username']

                                if follower_name not in home_followers:
                                    if follower_name not in all_followers:
                                        while not api.searchUsername(follower_name) or 'user' not in api.LastJson:
                                            if api.LastJson['message'] == 'Please wait a few minutes before you try again.':
                                                print('sleeping for a minute')
                                                time.sleep(60)
                                            else:
                                                break

                                        if 'user' in api.LastJson and api.LastJson['user']['follower_count'] > self.follower_limit:
                                            all_followers[follower_name] = {}
                                            all_followers[follower_name]['cnt'] = 1
                                            all_followers[follower_name]['follower_cnt'] = api.LastJson['user']['follower_count']
                                            all_followers[follower_name]['following_cnt'] = api.LastJson['user']['following_count']
                                            follower_sheet.cell(row=ind, column=1).value = user['name']
                                            follower_sheet.cell(row=ind, column=2).value = follower_name
                                            follower_sheet.cell(row=ind, column=3).value = all_followers[follower_name]['follower_cnt']
                                            follower_sheet.cell(row=ind, column=4).value = all_followers[follower_name]['following_cnt']
                                            ind += 1
                                    else:
                                        all_followers[follower_name]['cnt'] += 1
                                        follower_sheet.cell(row=ind, column=1).value = user['name']
                                        follower_sheet.cell(row=ind, column=2).value = follower_name
                                        follower_sheet.cell(row=ind, column=3).value = all_followers[follower_name]['follower_cnt']
                                        follower_sheet.cell(row=ind, column=4).value = all_followers[follower_name]['following_cnt']
                                        ind += 1
                                user['worked_count'] += 1
                                self.update_progress()

                        if not has_max_id:
                            break
                    except Exception as ex:
                        print(ex)
                        print('sleeping for a minute')
                        time.sleep(60)

            if 'working' in user['status']:
                user['status'] = 'completed'
                self.update_progress()

        all_followers_list = []
        for follower_name, value in all_followers.items():
            temp = [follower_name, value['cnt'], value['follower_cnt'], value['following_cnt']]
            all_followers_list.append(temp)

        all_followers_list = sorted(all_followers_list, key=lambda x: (-x[1], -x[2]))

        ind = 2
        for value in all_followers_list:
            summary_sheet.cell(row=ind, column=1).value = value[0]
            summary_sheet.cell(row=ind, column=2).value = value[1]
            summary_sheet.cell(row=ind, column=3).value = value[2]
            summary_sheet.cell(row=ind, column=4).value = value[3]
            ind += 1

        book.save(output_file)
        self.thread = None
        self.button_text.set("Start")
        logout()


def main():  # args=None
    root = tk.Tk()
    # root.eval('tk::PlaceWindow %s center' % root.winfo_pathname(root.winfo_id()))
    my_gui = MainView(root)
    root.mainloop()


if __name__ == '__main__':
    main()
